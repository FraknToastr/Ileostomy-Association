#!/usr/bin/env python3
"""Extract stocktake workbook rows into a browser-friendly JS data file."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK = Path("STOCKTAKE PHARMA FOR 30 06 2026 FINAL COUNTED.xlsx")
OUTPUT = Path("stocktake-data.js")


def number(value: Any) -> float:
    if value is None:
        return 0.0
    return float(value)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def round_money(value: float) -> float:
    return round(value + 0.000000001, 2)


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True)
    sheet = workbook["Sheet1"]

    items: list[dict[str, Any]] = []
    section = "Pharma"
    segment = "PBS"
    subsection = "Pharma items"
    # The workbook's first header is row 1, but extraction starts at row 2.
    # Start at block 1 so later repeated CODE headers become block 2, 3, ...
    block_index = 1

    for row_number in range(2, sheet.max_row + 1):
        row = [sheet.cell(row_number, col).value for col in range(1, 11)]
        code, description, company, price, sams, _total, actual, _diff, _diff_amount, _total_value = row

        if code == "CODE":
            block_index += 1
            continue

        if isinstance(code, str) and code.strip().upper() == "PAY FOR ITEMS":
            segment = "Non-PBS"
            section = "Pay for items"
            subsection = "Pay for items"
            continue

        if clean(description).upper() == "BASES/BAGS STOCK":
            segment = "PBS"
            section = "Bases/Bags stock"
            subsection = "Unassigned"
            continue

        if code is None and isinstance(description, str):
            label = description.strip()
            if label and label.upper() not in {
                "TOTAL PAY FOR",
                "TOTAL PHARMA (NOT PAY FOR)",
                "PBS TOTAL",
                "NON PBS TOTALS",
                "TOTAL",
            }:
                subsection = label.title()
            continue

        if code is None or description in (None, "Description"):
            continue

        price_value = number(price)
        sams_count = int(number(sams))
        actual_count = int(number(actual))
        quantity_diff = actual_count - sams_count
        value_diff = price_value * quantity_diff
        workbook_quantity_diff = number(_diff)
        workbook_value_diff = number(_diff_amount)

        if row_number < 303:
            segment = "PBS"
            section = "Pharma"
            subsection = f"Pharma block {max(block_index, 1)}"

        items.append(
            {
                "id": f"r{row_number}",
                "row": row_number,
                "code": clean(code),
                "description": clean(description),
                "company": clean(company),
                "section": section,
                "subsection": subsection,
                "segment": segment,
                "price": round_money(price_value),
                "sams": sams_count,
                "actual": actual_count,
                "quantityDiff": quantity_diff,
                "workbookQuantityDiff": int(workbook_quantity_diff),
                "samsValue": round_money(price_value * sams_count),
                "actualValue": round_money(price_value * actual_count),
                "valueDiff": round_money(value_diff),
                "workbookValueDiff": round_money(workbook_value_diff),
                "formulaIssue": round_money(value_diff) != round_money(workbook_value_diff),
            }
        )

    payload = {
        "source": WORKBOOK.name,
        "sheet": "Sheet1",
        "asAt": "2026-06-30",
        "generatedFromRows": len(items),
        "items": items,
    }

    OUTPUT.write_text(
        "window.STOCKTAKE_DATA = "
        + json.dumps(payload, ensure_ascii=True, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT} with {len(items)} items")


if __name__ == "__main__":
    main()
